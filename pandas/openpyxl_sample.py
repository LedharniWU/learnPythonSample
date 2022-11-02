import openpyxl
# ブック、シートを開く
wb = openpyxl.load_workbook("data.xlsx")

ws = wb[wb.sheetnames[0]]
c1 = ws.cell(row=2, column=2)
print(c1.value)